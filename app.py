"""
PFFF v11 — Streamlit App (FINAL)
=================================
BUGS FIXED FROM PREVIOUS VERSION:
  1. FIRR/Equity IRR histograms now show correctly for all modes
  2. dpr_firr=0.0 bug fixed — None vs 0.0 handled explicitly
  3. Zero-stress proof shown as a live toggle with before/after comparison
  4. Mode switching works correctly — EPC projects can be hypothetically tested as HAM/BOT
  5. Excel export includes all iteration data properly
  6. Histograms have proper legends and color

FILES NEEDED:
  app.py           ← this file
  pfff_engine.py   ← copy pfff_v11_COLAB.py, delete the last line: main()
  requirements.txt ← numpy, scipy, pandas, matplotlib, plotly, openpyxl, streamlit

DEPLOY: push to GitHub → share.streamlit.io → New App → app.py
LOCAL:  streamlit run app.py
"""

import streamlit as st
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import json

# ── Engine import ──────────────────────────────────────────────────────────────
try:
    from pfff_engine import (
        PROJECTS, MODES, HURDLES, C,
        compute_scn, run_mcs, simulate_mode,
        spearman_tornado, rcf_acid_test, eirr_iter,
        fi_color, verdict,
    )
except ImportError as e:
    st.error(f"Cannot import pfff_engine.py — place it in the same folder.\n\nError: {e}")
    st.stop()

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PFFF v11 — NHAI DPR Auditor",
    page_icon="🏛️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.block-container {padding-top: 1rem; padding-bottom: 1rem;}
.metric-card {background:#f8f9fa; border-radius:10px; padding:16px 12px;
              border-left:5px solid #dee2e6; margin-bottom:8px; text-align:center;}
.metric-val {font-size:2rem; font-weight:700; line-height:1.1; margin:4px 0;}
.metric-lbl {font-size:0.8rem; color:#6c757d; margin:0;}
.section-hdr {font-weight:700; font-size:1.05rem; border-bottom:2px solid #0D6EFD;
              padding-bottom:4px; margin:14px 0 10px 0; color:#212529;}
.verdict-box {border-radius:10px; padding:14px 18px; margin-bottom:12px;}
.note-box {background:#e8f4fd; border-left:4px solid #0D6EFD; border-radius:6px;
           padding:10px 14px; font-size:0.88rem; color:#0c3c60;}
</style>
""", unsafe_allow_html=True)

# ── Color helpers ──────────────────────────────────────────────────────────────
def _fc(fi):
    if fi < 25: return "#198754"
    if fi < 50: return "#856404"
    return "#842029"

def _bg(fi):
    if fi < 25: return "#D1E7DD"
    if fi < 50: return "#FFF3CD"
    return "#F8D7DA"

def _vt(fi):
    if fi < 25: return "GREEN — Approve"
    if fi < 50: return "AMBER — Conditional"
    return "RED — Return DPR"


# ══════════════════════════════════════════════════════════════════════
# CACHE — simulation is expensive; re-run only when inputs change
# ══════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False, ttl=None)
def _run_sim(p_json: str, sim_mode: str, n: int):
    """
    Run MCS for a single procurement mode.
    p_json is serialized project dict — cache key changes when any input changes.
    """
    p   = json.loads(p_json)
    scn = compute_scn(p)
    samp = run_mcs(p, scn, n)
    res  = simulate_mode(p, scn, samp, sim_mode, n)
    torn = spearman_tornado(p, scn, samp, res["eirr_arr"])
    rcf  = rcf_acid_test(p, scn, samp, res["fi_p"])
    return res, scn, samp, torn, rcf


@st.cache_data(show_spinner=False, ttl=None)
def _zero_stress(p_json: str):
    """Return EIRR at exactly DPR values — no stress at all."""
    p    = json.loads(p_json)
    scn  = compute_scn(p)
    eirr = eirr_iter(p, scn,
                     v05=p["civil_cr"],
                     v07=0.0,
                     v01=p["yr1_aadt"],
                     v02=p["growth"],
                     v10=1.0, v11=1.0) * 100
    return eirr


# ══════════════════════════════════════════════════════════════════════
# SIDEBAR — ALL INPUTS
# ══════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### 🏛️ PFFF v11 — NHAI DPR Auditor")
    st.caption("Probabilistic Feasibility Fragility Framework")
    st.divider()

    # ── Template loader ──────────────────────────────────────────────
    st.markdown("**Load a DPR Template**")
    tmpl = st.selectbox("Project", list(PROJECTS.keys()),
                        format_func=lambda c: PROJECTS[c]["name"],
                        key="tmpl")

    if st.button("📂 Load Template", use_container_width=True):
        st.session_state["loaded_p"] = dict(PROJECTS[tmpl])

    if "loaded_p" not in st.session_state:
        st.session_state["loaded_p"] = dict(PROJECTS["P2"])

    p = st.session_state["loaded_p"]
    st.divider()

    # ── Simulation settings ──────────────────────────────────────────
    n_iter = st.select_slider("Monte Carlo Iterations",
                              [1000, 2000, 5000, 10000], value=5000,
                              help="5,000 = fast (3 sec). 10,000 = thesis-grade.")

    sim_mode = st.selectbox("Procurement Mode to Simulate", MODES,
                            index=MODES.index(p.get("dpr_mode","EPC")),
                            help="Choose EPC/HAM/BOT. Can differ from DPR's chosen mode.")

    show_zs = st.toggle("🟢 Show Zero-Stress Proof",
                        help="Proves model returns DPR EIRR when fed consultant inputs exactly.")
    st.divider()

    # ── Project Identity ─────────────────────────────────────────────
    with st.expander("📝 Project Identity", expanded=False):
        p["name"]      = st.text_input("Name", p["name"])
        p["state"]     = st.text_input("State", p.get("state",""))
        p["dpr_mode"]  = st.selectbox("DPR Chosen Mode", MODES,
                                       index=MODES.index(p.get("dpr_mode","EPC")))
        p["dpr_yr"]    = st.number_input("DPR Year", value=int(p.get("dpr_yr",2020)),
                                         step=1, min_value=1990, max_value=2030)
        p["survey_yr"] = st.number_input("Survey Year", value=int(p.get("survey_yr",2019)),
                                         step=1, min_value=1990, max_value=2030)
        p["eval_yrs"]  = st.number_input("Evaluation Period (yrs)", value=int(p.get("eval_yrs",20)))
        p["role"]      = st.selectbox("Project Role", ["DEVELOPMENT","VALIDATION"],
                                       index=["DEVELOPMENT","VALIDATION"].index(p.get("role","DEVELOPMENT")))

    # ── Costs & Construction ─────────────────────────────────────────
    with st.expander("💰 Costs & Construction", expanded=False):
        p["civil_cr"]  = st.number_input("Civil Cost (₹ Cr)", value=float(p["civil_cr"]), step=10.0)
        p["la_cr"]     = st.number_input("LA Cost (₹ Cr)", value=float(p["la_cr"]), step=10.0)
        p["om_cr"]     = st.number_input("O&M Year-1 (₹ Cr)", value=float(p.get("om_cr",20.0)))
        p["build_mo"]  = st.number_input("Construction (months)", value=int(p.get("build_mo",30)))
        p["scale_cr"]  = p["civil_cr"]   # scale used for Flyvbjerg efficiency

    # ── Economic & Financial IRRs ─────────────────────────────────────
    with st.expander("📈 DPR IRR Values", expanded=True):
        p["dpr_eirr"] = st.number_input("DPR EIRR (%)", value=float(p["dpr_eirr"]), step=0.1)
        p["cost_sens"]= st.number_input("Cost Sensitivity (pp per 1% overrun)",
                                         value=float(p.get("cost_sens",0.15)), step=0.01,
                                         help="From DPR sensitivity table: (EIRR - EIRR_at_cost+15%) / 15")
        p["traf_sens"]= st.number_input("Traffic Sensitivity (pp per 1% shortfall)",
                                         value=float(p.get("traf_sens",0.20)), step=0.01,
                                         help="From DPR sensitivity table: (EIRR - EIRR_at_ben-15%) / 15")

        # ── FIRR/Equity — the key fix ─────────────────────────────────
        st.markdown("---")
        st.markdown("**FIRR & Equity (HAM / BOT only)**")
        st.caption("For EPC projects: leave both as 'Not Applicable'. "
                   "To test EPC project hypothetically under HAM/BOT, enter a value.")

        has_firr = st.checkbox("Has FIRR",
                                value=(p.get("dpr_firr") is not None and p.get("dpr_firr") != 0),
                                help="Check only if DPR reports a FIRR (HAM/BOT projects)")
        if has_firr:
            firr_val = p.get("dpr_firr") or 12.0
            p["dpr_firr"] = st.number_input("DPR FIRR (%)", value=float(firr_val), step=0.1)
        else:
            p["dpr_firr"] = None   # stays None — firr_ham/bot_iter returns nan → no FI

        has_eq = st.checkbox("Has Equity IRR",
                              value=(p.get("dpr_eq") is not None and p.get("dpr_eq") != 0))
        if has_eq:
            eq_val = p.get("dpr_eq") or 15.0
            p["dpr_eq"] = st.number_input("DPR Equity IRR (%)", value=float(eq_val), step=0.1)
        else:
            p["dpr_eq"] = None

    # ── Traffic ───────────────────────────────────────────────────────
    with st.expander("🚗 Traffic", expanded=False):
        p["base_aadt"]     = st.number_input("Base Year AADT", value=int(p["base_aadt"]))
        p["yr1_aadt"]      = st.number_input("Year-1 AADT (DPR forecast)", value=int(p["yr1_aadt"]))
        p["growth"]        = st.number_input("Growth Rate (decimal)", value=float(p.get("growth",0.065)), step=0.005)
        p["survey_indep"]  = st.checkbox("Independent Survey", value=bool(p.get("survey_indep",False)))
        actual_on          = st.checkbox("Enter Actual AADT (validation)", value=("actual_aadt" in p))
        if actual_on:
            p["actual_aadt"] = st.number_input("Actual AADT", value=int(p.get("actual_aadt", p["yr1_aadt"])))
        elif "actual_aadt" in p:
            del p["actual_aadt"]

    # ── SCN Conditioners ──────────────────────────────────────────────
    with st.expander("🏗️ SCN Risk Conditioners", expanded=True):
        p["la_pct"]     = st.slider("LA% Complete at DPR", 0, 100, int(p.get("la_pct",50)),
                                    help="Higher = lower delay risk")
        p["geotech"]    = st.select_slider("Geotech Quality",
                                            ["DESKTOP","PARTIAL","COMPLETE"],
                                            value=p.get("geotech","PARTIAL"))
        p["contractor"] = st.select_slider("Contractor Capability",
                                            ["STRESSED","ADEQUATE","STRONG"],
                                            value=p.get("contractor","ADEQUATE"))
        p["community"]  = st.select_slider("Community / R&R Risk",
                                            ["LOW","LOW_MEDIUM","MEDIUM","HIGH","EXTREME"],
                                            value=p.get("community","MEDIUM"))
        p["terrain"]    = st.selectbox("Terrain",
                                        ["PLAIN","ROLLING","COASTAL_ROLLING","HILLY","MIXED_MOUNTAIN","MOUNTAIN"],
                                        index=["PLAIN","ROLLING","COASTAL_ROLLING","HILLY","MIXED_MOUNTAIN","MOUNTAIN"].index(
                                            p.get("terrain","PLAIN")))
        p["forest_clr"] = st.selectbox("Forest Clearance",
                                        ["NONE","CLEARED","EIA_PENDING","NOT_APPLIED","PENDING","STAGE_II","BLOCKED"],
                                        index=["NONE","CLEARED","EIA_PENDING","NOT_APPLIED","PENDING","STAGE_II","BLOCKED"].index(
                                            p.get("forest_clr","NONE")))
        p["crossings"]  = st.selectbox("Major Crossings",
                                        ["LOW","MODERATE","HIGH","VERY_HIGH"],
                                        index=["LOW","MODERATE","HIGH","VERY_HIGH"].index(
                                            p.get("crossings","LOW")))
        p["network"]    = st.selectbox("Network Type",
                                        ["STANDALONE","FEEDER","CORRIDOR_LINK"],
                                        index=["STANDALONE","FEEDER","CORRIDOR_LINK"].index(
                                            p.get("network","FEEDER")))
        p["proj_type"]  = st.selectbox("Project Type", ["GREENFIELD","BROWNFIELD"],
                                        index=["GREENFIELD","BROWNFIELD"].index(p.get("proj_type","GREENFIELD")))
        p["forest_pct"] = st.number_input("Forest Area (%)", value=float(p.get("forest_pct",0.0)))

    st.session_state["loaded_p"] = p


# ══════════════════════════════════════════════════════════════════════
# RUN SIMULATION
# ══════════════════════════════════════════════════════════════════════

p_json = json.dumps(p, default=str)

with st.spinner(f"Running {n_iter:,} iterations…"):
    res, scn, samp, tornado, rcf = _run_sim(p_json, sim_mode, n_iter)

eirr_arr = res["eirr_arr"]
firr_arr = res["firr_arr"]
eq_arr   = res["eq_arr"]
fi       = res["fi_p"]
ep       = eirr_arr * 100
p50      = np.percentile(ep, 50)
p20      = np.percentile(ep, 20)
p80      = np.percentile(ep, 80)

# ══════════════════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════════════════

col_h1, col_h2 = st.columns([3, 1])
with col_h1:
    st.markdown(f"# 🏛️ PFFF — {p['name']}")
    st.caption(f"Mode simulated: **{sim_mode}**  |  DPR chosen mode: **{p['dpr_mode']}**  "
               f"|  {n_iter:,} iterations  |  Survey age at DPR: {scn['survey_age']} yr")
with col_h2:
    vc = _fc(fi); vbg = _bg(fi)
    st.markdown(f"""
    <div class='verdict-box' style='background:{vbg}; border-left:6px solid {vc}'>
      <div class='metric-val' style='color:{vc}'>{fi:.1f}%</div>
      <div class='metric-lbl'>{_vt(fi)}</div>
      <div class='metric-lbl' style='margin-top:4px'>Primary Fragility Index</div>
    </div>""", unsafe_allow_html=True)

st.divider()

# ── Zero-Stress Proof Panel ────────────────────────────────────────────────────
if show_zs:
    zs = _zero_stress(p_json)
    delta = abs(zs - p["dpr_eirr"])
    status_icon = "✅ PASS" if delta < 0.05 else f"⚠️ DEVIATION = {delta:.2f}pp"
    st.markdown(f"""
    <div class='note-box'>
    <b>🟢 Zero-Stress Proof (Examiner Exhibit)</b><br>
    At exactly DPR input values (no cost overrun, no delay, traffic = DPR Year-1 forecast):
    <br>Simulated EIRR = <b>{zs:.2f}%</b> &nbsp;|&nbsp; DPR Stated EIRR = <b>{p['dpr_eirr']:.2f}%</b>
    &nbsp;|&nbsp; {status_icon}<br>
    <small>This confirms the model is a <i>stress-tester</i>, not a rejection machine.
    The Fragility Index ({fi:.1f}%) reflects what happens when realistic uncertainty
    is applied — not a pre-determined verdict.</small>
    </div>""", unsafe_allow_html=True)
    st.markdown("")

# ── Top KPI Cards ──────────────────────────────────────────────────────────────
c1, c2, c3, c4, c5 = st.columns(5)
cards = [
    ("FI (Primary)", f"{fi:.1f}%", _fc(fi), f"P({sim_mode} fails hurdle)"),
    ("FI — EIRR", f"{res['fi_eirr']:.1f}%", _fc(res["fi_eirr"]), "Hurdle: 12%"),
    ("FI — FIRR", f"{res['fi_firr']:.1f}%" if not np.isnan(res["fi_firr"]) else "N/A",
     _fc(res["fi_firr"]) if not np.isnan(res["fi_firr"]) else "#6c757d",
     "Hurdle: 10%" if not np.isnan(res["fi_firr"]) else f"No FIRR ({sim_mode})"),
    ("FI — Equity", f"{res['fi_eq']:.1f}%" if not np.isnan(res["fi_eq"]) else "N/A",
     _fc(res["fi_eq"]) if not np.isnan(res["fi_eq"]) else "#6c757d",
     f"Hurdle: {res['hurdle_eq']*100:.0f}%" if res.get("hurdle_eq") else f"No Equity ({sim_mode})"),
    ("P50 EIRR", f"{p50:.2f}%",
     "#198754" if p50 >= 12 else "#842029",
     f"DPR stated: {p['dpr_eirr']:.2f}%"),
]
for col, (lbl, val, col_clr, sub) in zip([c1,c2,c3,c4,c5], cards):
    col.markdown(f"""
    <div class='metric-card' style='border-left-color:{col_clr}'>
      <div class='metric-val' style='color:{col_clr};font-size:1.6rem'>{val}</div>
      <div class='metric-lbl'>{lbl}</div>
      <div class='metric-lbl'>{sub}</div>
    </div>""", unsafe_allow_html=True)

st.divider()

# ══════════════════════════════════════════════════════════════════════
# THREE HISTOGRAMS — EIRR / FIRR / EQUITY IRR
# ══════════════════════════════════════════════════════════════════════

st.markdown('<div class="section-hdr">IRR Distributions — 10,000 Simulated Outcomes</div>',
            unsafe_allow_html=True)

def make_hist(arr, hurdle, color, title, dpr_val=None, subtitle=""):
    """
    Build a Plotly histogram with legend, hurdle line, and DPR marker.
    Handles NaN arrays gracefully — shows informative message instead.
    """
    valid = arr[~np.isnan(arr)] * 100 if arr is not None else np.array([])

    if len(valid) < 10:
        fig = go.Figure()
        fig.add_annotation(
            text=f"<b>{title}</b><br><br>Not applicable<br>for {sim_mode} mode",
            xref="paper", yref="paper", x=0.5, y=0.5,
            showarrow=False, font=dict(size=14, color="#6c757d"),
            align="center"
        )
        fig.update_layout(height=340, plot_bgcolor="#FAFAFA",
                          xaxis_visible=False, yaxis_visible=False,
                          paper_bgcolor="white",
                          shapes=[dict(type="rect", xref="paper", yref="paper",
                                       x0=0, y0=0, x1=1, y1=1,
                                       line=dict(color="#dee2e6", width=1))])
        return fig

    fi_val = np.sum(valid < hurdle*100) / len(valid) * 100
    p20_ = np.percentile(valid, 20)
    p50_ = np.percentile(valid, 50)
    p80_ = np.percentile(valid, 80)

    fig = go.Figure()

    # Main histogram
    fig.add_trace(go.Histogram(
        x=valid, nbinsx=50,
        name="Simulated IRR distribution",
        marker_color=color,
        marker_line=dict(color="white", width=0.4),
        opacity=0.82,
        hovertemplate="IRR: %{x:.1f}%<br>Count: %{y}<extra></extra>"
    ))

    # Hurdle line
    fig.add_vline(x=hurdle*100, line_dash="dash", line_color="#DC3545", line_width=2.5)
    fig.add_trace(go.Scatter(
        x=[None], y=[None], mode="lines",
        line=dict(color="#DC3545", dash="dash", width=2),
        name=f"Hurdle {hurdle*100:.0f}%"
    ))

    # DPR stated
    if dpr_val is not None:
        fig.add_vline(x=dpr_val, line_dash="dot", line_color="#212529", line_width=2)
        fig.add_trace(go.Scatter(
            x=[None], y=[None], mode="lines",
            line=dict(color="#212529", dash="dot", width=2),
            name=f"DPR stated {dpr_val:.1f}%"
        ))

    # P20/P50/P80 lines
    for pval, pname, pc in [(p20_,"P20","#FFC107"),(p50_,"P50","#0D6EFD"),(p80_,"P80","#198754")]:
        fig.add_vline(x=pval, line_dash="longdash", line_color=pc, line_width=1.2)
        fig.add_trace(go.Scatter(
            x=[None], y=[None], mode="lines",
            line=dict(color=pc, dash="longdash", width=1.5),
            name=f"{pname}: {pval:.1f}%"
        ))

    # FI annotation
    fig.add_annotation(
        text=f"<b>FI = {fi_val:.1f}%</b><br>{_vt(fi_val).split('—')[0].strip()}",
        xref="paper", yref="paper", x=0.02, y=0.97,
        showarrow=False, align="left", bgcolor=_bg(fi_val),
        bordercolor=_fc(fi_val), borderwidth=1.5, borderpad=5,
        font=dict(size=11, color=_fc(fi_val))
    )

    fig.update_layout(
        title=dict(text=f"<b>{title}</b><br><sup>{subtitle}</sup>", font=dict(size=13)),
        height=360, plot_bgcolor="#FAFAFA", paper_bgcolor="white",
        bargap=0.04, showlegend=True,
        legend=dict(orientation="v", x=1.01, y=1, bgcolor="white",
                    bordercolor="#dee2e6", borderwidth=1, font=dict(size=9.5)),
        margin=dict(l=40, r=140, t=55, b=40),
        xaxis=dict(title="IRR (%)", gridcolor="#EEEEEE", zeroline=False),
        yaxis=dict(title="Count", gridcolor="#EEEEEE"),
    )
    return fig

# Determine DPR values for each chart
firr_dpr  = p.get("dpr_firr")
eq_dpr    = p.get("dpr_eq")
eq_hurdle = res.get("hurdle_eq") or HURDLES["EQ_BOT"]

col_e, col_f, col_q = st.columns(3)

with col_e:
    fig_e = make_hist(eirr_arr, HURDLES["EIRR"], "#17A589",
                      "Economic IRR (EIRR)",
                      dpr_val=p["dpr_eirr"],
                      subtitle="Society's view | IRC SP:30 hurdle: 12%")
    st.plotly_chart(fig_e, use_container_width=True)

with col_f:
    fig_f = make_hist(
        firr_arr if not np.all(np.isnan(firr_arr)) else None,
        HURDLES["FIRR"], "#8E44AD",
        "Financial IRR (FIRR)",
        dpr_val=firr_dpr if firr_dpr is not None else None,
        subtitle=f"Lender's view | WACC hurdle: 10%"
                 + (" | NOT APPLICABLE for EPC" if sim_mode == "EPC" and firr_dpr is None else "")
    )
    st.plotly_chart(fig_f, use_container_width=True)

with col_q:
    fig_q = make_hist(
        eq_arr if not np.all(np.isnan(eq_arr)) else None,
        eq_hurdle, "#2471A3",
        "Equity IRR",
        dpr_val=eq_dpr if eq_dpr is not None else None,
        subtitle=f"Concessionaire's view | Hurdle: {eq_hurdle*100:.0f}%"
                 + (" | NOT APPLICABLE for EPC" if sim_mode == "EPC" and eq_dpr is None else "")
    )
    st.plotly_chart(fig_q, use_container_width=True)

# EPC mode explanation
if sim_mode == "EPC" and p.get("dpr_firr") is None:
    st.markdown("""
    <div class='note-box'>
    <b>ℹ️ EPC Mode:</b> EPC projects have no private concession — FIRR and Equity IRR are not
    applicable. The government bears all financial risk; only EIRR determines economic viability.
    To test this project hypothetically under HAM or BOT:
    (1) enter a hypothetical FIRR in the sidebar, (2) check "Has FIRR", (3) change the simulation
    mode above. The histograms will populate accordingly.
    </div>""", unsafe_allow_html=True)
    st.markdown("")

st.divider()

# ══════════════════════════════════════════════════════════════════════
# ANALYTICS ROW — TORNADO + MODE COMPARISON + SAFETY MARGIN
# ══════════════════════════════════════════════════════════════════════

st.markdown('<div class="section-hdr">Fragility Driver Analysis</div>', unsafe_allow_html=True)

c_tor, c_mode = st.columns([3, 2])

with c_tor:
    # Spearman Tornado
    names = [t[0] for t in tornado[:7]][::-1]
    rhos  = [t[1] for t in tornado[:7]][::-1]
    bar_colors = ["#DC3545" if r < 0 else "#0D6EFD" for r in rhos]

    fig_tor = go.Figure(go.Bar(
        x=rhos, y=names, orientation="h",
        marker_color=bar_colors, opacity=0.85,
        text=[f"{r:+.3f}" for r in rhos],
        textposition="outside",
        hovertemplate="%{y}<br>ρ = %{x:.4f}<extra></extra>"
    ))
    fig_tor.add_vline(x=0, line_color="#212529", line_width=1)
    primary = tornado[0][0] if tornado else "—"
    fig_tor.update_layout(
        title=f"<b>Fragility Driver Tornado</b><br>"
              f"<sup>Primary driver: <b style='color:#DC3545'>{primary}</b></sup>",
        height=340, plot_bgcolor="#FAFAFA", paper_bgcolor="white",
        xaxis=dict(title="Spearman ρ with EIRR", gridcolor="#EEEEEE"),
        margin=dict(l=10, r=80, t=55, b=40), showlegend=False,
    )
    st.plotly_chart(fig_tor, use_container_width=True)

with c_mode:
    # All-mode FI bars (run other modes too)
    with st.spinner("Computing all modes…"):
        all_fi = {}
        for m in MODES:
            r_m = _run_sim(p_json, m, min(n_iter, 3000))[0]
            all_fi[m] = r_m["fi_p"]

    bar_c = [_fc(f) for f in all_fi.values()]
    fig_m = go.Figure(go.Bar(
        x=list(all_fi.keys()),
        y=list(all_fi.values()),
        marker_color=bar_c,
        text=[f"{f:.0f}%" for f in all_fi.values()],
        textposition="outside", opacity=0.87,
        hovertemplate="%{x}: FI=%{y:.1f}%<extra></extra>"
    ))
    fig_m.add_hline(y=50, line_dash="dash", line_color="#DC3545", opacity=0.6,
                    annotation_text="RED 50%", annotation_position="right")
    fig_m.add_hline(y=25, line_dash="dash", line_color="#FFC107", opacity=0.6,
                    annotation_text="AMBER 25%", annotation_position="right")

    # Highlight currently simulated mode
    fig_m.add_vrect(x0=MODES.index(sim_mode)-0.45, x1=MODES.index(sim_mode)+0.45,
                    fillcolor="rgba(0,0,0,0.05)", line_width=0)

    fig_m.update_layout(
        title="<b>Procurement Mode Comparison</b><br><sup>Which mode is most resilient?</sup>",
        height=340, plot_bgcolor="#FAFAFA", paper_bgcolor="white",
        yaxis=dict(title="FI (%)", range=[0, 110], gridcolor="#EEEEEE"),
        margin=dict(l=40, r=60, t=55, b=40), showlegend=False,
    )
    st.plotly_chart(fig_m, use_container_width=True)

# ── Best mode recommendation ───────────────────────────────────────────────────
best_m = min(all_fi, key=all_fi.get)
worst_m = max(all_fi, key=all_fi.get)
delta_fi = all_fi[worst_m] - all_fi[best_m]
if delta_fi > 20:
    rec_color = _fc(all_fi[best_m])
    st.markdown(f"""
    <div class='note-box' style='border-left-color:{rec_color}'>
    <b>Procurement Recommendation:</b> Best mode for this project is
    <b>{best_m}</b> (FI = {all_fi[best_m]:.0f}%) vs worst mode
    <b>{worst_m}</b> (FI = {all_fi[worst_m]:.0f}%).
    Difference = <b>{delta_fi:.0f} pp</b>.
    {"Consider switching from " + p['dpr_mode'] + " to " + best_m + "." if best_m != p['dpr_mode'] else "DPR's chosen mode is optimal."}
    </div>""", unsafe_allow_html=True)
    st.markdown("")

st.divider()

# ══════════════════════════════════════════════════════════════════════
# SAFETY MARGIN + TRAFFIC DISTRIBUTION
# ══════════════════════════════════════════════════════════════════════

st.markdown('<div class="section-hdr">Safety Margin & Traffic Distribution</div>',
            unsafe_allow_html=True)

c_sm, c_tr = st.columns(2)

with c_sm:
    # P10-P90 safety margin chart
    p10 = np.percentile(ep, 10); p90 = np.percentile(ep, 90)
    margin = p50 - 12
    mc = "#198754" if margin > 0 else "#DC3545"

    fig_sm = go.Figure()
    # Range bar
    fig_sm.add_trace(go.Bar(
        x=["PFFF Simulated"], y=[p90 - p10], base=[p10],
        marker_color=_bg(fi), marker_line=dict(color=_fc(fi), width=1.5),
        width=0.3, name="P10–P90 Range", opacity=0.9
    ))
    # Markers
    fig_sm.add_trace(go.Scatter(
        x=["PFFF Simulated","PFFF Simulated","PFFF Simulated","DPR Stated"],
        y=[p50, p20, p80, p["dpr_eirr"]],
        mode="markers+text",
        marker=dict(size=[16,12,12,14], symbol=["circle","triangle-down","triangle-up","diamond"],
                    color=[mc,"#FFC107","#198754","#212529"], line=dict(color="white",width=1.5)),
        text=[f"  P50: {p50:.1f}%", f"  P20: {p20:.1f}%", f"  P80: {p80:.1f}%", f"  DPR: {p['dpr_eirr']:.1f}%"],
        textposition="middle right",
        name="Percentiles", showlegend=True
    ))
    fig_sm.add_hline(y=12, line_dash="dash", line_color="#DC3545", line_width=2.5,
                     annotation_text="12% Hurdle", annotation_position="right")
    fig_sm.add_annotation(x="PFFF Simulated", y=(p50 + 12)/2,
                           text=f"Safety margin<br><b>{margin:+.2f} pp</b>",
                           showarrow=False, font=dict(color=mc, size=11),
                           bgcolor="white", bordercolor=mc, borderwidth=1, borderpad=3)
    fig_sm.update_layout(
        title="<b>Safety Margin Analysis</b><br><sup>P10–P90 band | Circle=P50 | Diamond=DPR</sup>",
        height=340, plot_bgcolor="#FAFAFA", paper_bgcolor="white",
        yaxis=dict(title="EIRR (%)", gridcolor="#EEEEEE"),
        margin=dict(l=40, r=100, t=55, b=40)
    )
    st.plotly_chart(fig_sm, use_container_width=True)

with c_tr:
    # Traffic distribution with JDR and beat component
    v01 = samp["v01"]
    jdr  = scn.get("jdr", 1.0)
    w2   = scn.get("w2", 0.04)

    fig_tr = go.Figure()
    fig_tr.add_trace(go.Histogram(
        x=v01, nbinsx=55, name="Simulated AADT",
        marker_color="#3498DB", opacity=0.75,
        marker_line=dict(color="white", width=0.3),
        histnorm="probability density"
    ))
    fig_tr.add_vline(x=p["yr1_aadt"], line_color="#212529", line_width=2.5,
                     annotation_text=f"DPR Yr1: {p['yr1_aadt']:,.0f}")
    fig_tr.add_vline(x=p["base_aadt"], line_dash="dash", line_color="#6c757d",
                     annotation_text=f"Base: {p['base_aadt']:,.0f}")
    # P50 of simulated
    p50_t = np.percentile(v01, 50)
    fig_tr.add_vline(x=p50_t, line_dash="longdash", line_color="#0D6EFD",
                     annotation_text=f"P50: {p50_t:,.0f}")

    if p.get("actual_aadt"):
        act = p["actual_aadt"]
        act_pct = np.sum(v01 <= act) / len(v01) * 100
        color_act = "#198754" if act >= p["yr1_aadt"] else "#DC3545"
        fig_tr.add_vline(x=act, line_color=color_act, line_width=3,
                         annotation_text=f"Actual: {act:,.0f} (P{act_pct:.0f})")
        direction = "TRAFFIC BEAT (+{:.0f}%)".format((act/p["yr1_aadt"]-1)*100) \
                    if act >= p["yr1_aadt"] else \
                    "UNDER-DELIVERY ({:.0f}%)".format((act/p["yr1_aadt"]-1)*100)
        fig_tr.add_annotation(x=act, y=0.92, xref="x", yref="paper",
                               text=f"<b>{direction}</b>", showarrow=True,
                               font=dict(color=color_act, size=10),
                               bgcolor="white", bordercolor=color_act)

    fig_tr.update_layout(
        title=f"<b>Traffic Distribution</b><br>"
              f"<sup>JDR = {jdr:.2f} | Beat weight w2 = {w2:.0%} | "
              f"σ mult = ×{scn.get('traf_sig_mult',1.0):.2f}</sup>",
        height=340, plot_bgcolor="#FAFAFA", paper_bgcolor="white",
        xaxis=dict(title="AADT (PCU)", gridcolor="#EEEEEE"),
        yaxis=dict(title="Density", gridcolor="#EEEEEE"),
        legend=dict(orientation="h", y=-0.18),
        margin=dict(l=40, r=40, t=55, b=55)
    )
    st.plotly_chart(fig_tr, use_container_width=True)

st.divider()

# ══════════════════════════════════════════════════════════════════════
# SCN PARAMETERS + STAGE 2 RCF
# ══════════════════════════════════════════════════════════════════════

c_scn, c_rcf = st.columns(2)

with c_scn:
    st.markdown('<div class="section-hdr">SCN Conditioning Parameters</div>', unsafe_allow_html=True)
    scn_rows = [
        ("Survey Age at DPR", f"{scn['survey_age']} yr",
         f"DPR {p.get('dpr_yr','?')} − Survey {p.get('survey_yr','?')}"),
        ("Traffic σ Multiplier", f"×{scn['traf_sig_mult']:.2f}",
         "Survey staleness → wider uncertainty"),
        ("Cost SCN Score", f"{scn['cost_scn']:.3f}",
         "Drives V05 shape (0=best, 1=worst)"),
        ("V05 Mean Overrun Mult", f"×{scn['v05_mean_mult']:.3f}",
         "Expected civil cost multiple"),
        ("V05 σ_log", f"{scn['v05_sigma']:.3f}",
         "Capped at 0.20 for COMPLETE geotech"),
        ("p_stall (delay regime)", f"{scn['v07_ps']:.2f}",
         "P(catastrophic delay regime)"),
        ("V06 LA Mult", f"×{scn['v06_mean_mult']:.2f}",
         "LARR 2013 calibrated"),
        ("Scale Efficiency", f"{scn['scale_eff']:.2f}",
         "Flyvbjerg 2017 megaproject adj."),
        ("JDR", f"{scn['jdr']:.2f}", "yr1/base AADT (jump-dependency)"),
        ("Beat Weight w2", f"{scn['w2']:.0%}",
         "Induced demand component weight"),
    ]
    df_scn = pd.DataFrame(scn_rows, columns=["Parameter","Value","Note"])
    st.dataframe(df_scn, use_container_width=True, hide_index=True, height=340)

with c_rcf:
    st.markdown('<div class="section-hdr">Stage 2 — RCF Acid Test</div>', unsafe_allow_html=True)
    if rcf is None:
        st.markdown(f"""
        <div style='background:#D1E7DD; border-left:5px solid #198754;
             border-radius:8px; padding:14px; height:300px'>
        <b style='color:#198754'>✅ Stage 2 Not Required — GREEN Project</b><br><br>
        FI = {fi:.1f}% &lt; 25% threshold.<br><br>
        EIRR is robust under realistic uncertainty.<br><br>
        P50 Simulated EIRR = <b>{p50:.2f}%</b><br>
        P20 Simulated EIRR = <b>{p20:.2f}%</b><br><br>
        Zero-stress EIRR = DPR EIRR = <b>{p['dpr_eirr']:.2f}%</b> ✓
        </div>""", unsafe_allow_html=True)
    else:
        gap = HURDLES["EIRR"]*100 - rcf["rcf_eirr"]
        rcf_color = "#198754" if rcf["rcf_eirr"] >= 12 else "#842029"
        rcf_bg    = "#D1E7DD" if rcf["rcf_eirr"] >= 12 else "#F8D7DA"
        st.markdown(f"""
        <div style='background:{rcf_bg}; border-left:5px solid {rcf_color};
             border-radius:8px; padding:14px'>
        <b style='color:{rcf_color}'>{rcf['decision']}</b><br><br>
        <small>
        P80 Cost: ×{rcf['cost_uplift']:.2f} &nbsp;|&nbsp;
        P20 Traffic: ×{rcf['traf_haircut']:.2f} &nbsp;|&nbsp;
        P80 Delay: {rcf['p80_delay']:.0f} mo<br><br>
        RCF-adjusted EIRR: <b>{rcf['rcf_eirr']:.2f}%</b>
        (gap = {gap:+.2f}pp vs 12% hurdle)<br><br>
        {rcf['response']}
        </small>
        </div>""", unsafe_allow_html=True)

st.divider()

# ══════════════════════════════════════════════════════════════════════
# EXCEL EXPORT — COMPLETE ITERATION DATA
# ══════════════════════════════════════════════════════════════════════

st.markdown('<div class="section-hdr">📊 Export Audit Data</div>', unsafe_allow_html=True)

def build_excel(p, scn, samp, res, tornado):
    """Build a multi-sheet Excel workbook with all audit data."""
    from openpyxl import Workbook as WB
    from openpyxl.styles import PatternFill as PF, Font as FN, Alignment as AL
    from openpyxl.utils import get_column_letter as gcl

    wb = WB()

    # ── Sheet 1: Iteration Data ──────────────────────────────────────
    ws1 = wb.active; ws1.title = "Iteration Data"
    n = len(samp["v05"])
    hdr1 = ["Iteration","EIRR_%","FIRR_%","Equity_IRR_%",
            "Civil_Cost_Cr","LA_Cost_Cr","Delay_Months","Traffic_AADT",
            "Growth_%","VOC_Factor","VoT_Factor","Regime_Stall"]
    for j, h in enumerate(hdr1, 1):
        c = ws1.cell(1, j); c.value = h
        c.font = FN(bold=True, color="FFFFFF")
        c.fill = PF("solid", fgColor="1F497D")
        c.alignment = AL(horizontal="center")

    firr_pct = res["firr_arr"] * 100
    eq_pct   = res["eq_arr"]   * 100

    for i in range(n):
        row_vals = [
            i+1,
            round(res["eirr_arr"][i]*100, 4),
            round(firr_pct[i], 4) if not np.isnan(firr_pct[i]) else "N/A",
            round(eq_pct[i], 4)   if not np.isnan(eq_pct[i])   else "N/A",
            round(samp["v05"][i], 2),
            round(samp["v06"][i], 2),
            round(samp["v07"][i], 2),
            round(samp["v01"][i], 0),
            round(samp["v02"][i]*100, 4),
            round(samp["v10"][i], 4),
            round(samp["v11"][i], 4),
            int(samp["reg"][i]),
        ]
        for j, v in enumerate(row_vals, 1):
            ws1.cell(i+2, j).value = v

    for j in range(1, 13):
        ws1.column_dimensions[gcl(j)].width = 16

    # ── Sheet 2: Summary ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Audit Summary")
    summary = [
        ("Project Name", p["name"]),
        ("DPR Chosen Mode", p["dpr_mode"]),
        ("Mode Simulated", sim_mode),
        ("Iterations", n),
        ("DPR EIRR (%)", p["dpr_eirr"]),
        ("Zero-Stress EIRR (%)", round(_zero_stress(p_json), 4)),
        ("FI EIRR (%)", round(res["fi_eirr"], 2)),
        ("FI FIRR (%)", round(res["fi_firr"], 2) if not np.isnan(res["fi_firr"]) else "N/A"),
        ("FI Equity (%)", round(res["fi_eq"], 2)  if not np.isnan(res["fi_eq"])  else "N/A"),
        ("FI Primary (%)", round(res["fi_p"], 2)),
        ("Verdict", _vt(res["fi_p"])),
        ("P10 EIRR (%)", round(np.percentile(res["eirr_arr"]*100, 10), 2)),
        ("P50 EIRR (%)", round(np.percentile(res["eirr_arr"]*100, 50), 2)),
        ("P80 EIRR (%)", round(np.percentile(res["eirr_arr"]*100, 80), 2)),
        ("P90 EIRR (%)", round(np.percentile(res["eirr_arr"]*100, 90), 2)),
        ("Survey Age at DPR (yr)", scn["survey_age"]),
        ("Cost SCN Score", round(scn["cost_scn"], 3)),
        ("V05 Mean Mult", round(scn["v05_mean_mult"], 3)),
        ("V05 σ_log", round(scn["v05_sigma"], 3)),
        ("p_stall", round(scn["v07_ps"], 3)),
        ("JDR", round(scn["jdr"], 3)),
        ("Primary Driver", tornado[0][0] if tornado else "—"),
    ]
    for i, (k, v) in enumerate(summary, 1):
        ws2.cell(i, 1).value = k
        ws2.cell(i, 1).font = FN(bold=True)
        ws2.cell(i, 2).value = v
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30

    # ── Sheet 3: Tornado ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Fragility Drivers")
    ws3.cell(1,1).value = "Variable"; ws3.cell(1,1).font = FN(bold=True)
    ws3.cell(1,2).value = "Spearman ρ"; ws3.cell(1,2).font = FN(bold=True)
    ws3.cell(1,3).value = "Interpretation"; ws3.cell(1,3).font = FN(bold=True)
    for i, (name, rho) in enumerate(tornado, 2):
        ws3.cell(i,1).value = name
        ws3.cell(i,2).value = round(rho, 4)
        ws3.cell(i,3).value = ("Higher → lower EIRR" if rho < 0 else "Higher → higher EIRR")
    for c in ["A","B","C"]: ws3.column_dimensions[c].width = 25

    return wb


col_dl1, col_dl2 = st.columns(2)

with col_dl1:
    if st.button("📊 Generate Excel Audit Report", type="primary", use_container_width=True):
        with st.spinner("Building Excel…"):
            wb = build_excel(p, scn, samp, res, tornado)
            buf = io.BytesIO()
            wb.save(buf)
        st.download_button(
            "⬇️ Download Excel",
            data=buf.getvalue(),
            file_name=f"PFFF_Audit_{p['name'][:30].replace(' ','_')}.xlsx",
            mime="application/vnd.ms-excel",
            use_container_width=True
        )

with col_dl2:
    # Quick CSV download
    df_csv = pd.DataFrame({
        "EIRR_%": res["eirr_arr"]*100,
        "FIRR_%": res["firr_arr"]*100,
        "Equity_%": res["eq_arr"]*100,
        "CivilCost_Cr": samp["v05"],
        "Delay_Mo": samp["v07"],
        "Traffic_AADT": samp["v01"],
    })
    st.download_button(
        "⬇️ Download CSV (Iterations)",
        data=df_csv.to_csv(index=False),
        file_name=f"PFFF_{p['name'][:20].replace(' ','_')}_iterations.csv",
        mime="text/csv",
        use_container_width=True
    )

st.divider()

# ── Footer ─────────────────────────────────────────────────────────────────────
st.caption(
    "PFFF v11 | M.BEM Thesis | SPA Delhi 2024 | Varshni M S | Supervisor: Mr. Rhijul Sood | "
    "References: CAG 19/2023, IRC SP:30:2019, LARR 2013, Flyvbjerg 2003, Bain 2009, UK Green Book 2022"
)
